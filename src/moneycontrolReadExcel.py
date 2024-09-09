from openpyxl import Workbook
from openpyxl import load_workbook
from stopAds import main

try:
    wb = load_workbook('data.xlsx')
except FileNotFoundError:
    wb = Workbook()

ws = wb.active
url_list = [
    "https://www.moneycontrol.com/india/stockpricequote/power-generationdistribution/thetatapowercompany/TPC",
    "https://www.moneycontrol.com/india/stockpricequote/power-generationdistribution/adanipower/AP11",
    "https://www.moneycontrol.com/india/stockpricequote/infrastructure-general/adaniportsspecialeconomiczone/MPS",
    "https://www.moneycontrol.com/india/stockpricequote/metals-non-ferrous/hindustancopper/HC07",
    "https://www.moneycontrol.com/india/stockpricequote/constructioncontracting-civil/ncc/NCC01",
    "https://www.moneycontrol.com/india/stockpricequote/online-services/zomato/Z",
    "https://www.moneycontrol.com/india/stockpricequote/consumer-food/devyaniinternational/DI06",
    "https://www.moneycontrol.com/india/stockpricequote/engineering/engineersindia/EI14",
    "https://www.moneycontrol.com/india/stockpricequote/cement-mini/nclindustries/NCL",
    "https://www.moneycontrol.com/india/stockpricequote/power-generationdistribution/nhpc/N07",
    "https://www.moneycontrol.com/india/stockpricequote/power-generationdistribution/sjvn/S11",
    "https://www.moneycontrol.com/india/stockpricequote/diversified/tatatechnologies/TTL01",
    "https://www.moneycontrol.com/india/stockpricequote/engineering-industrial-equipments/ideaforgetechnology/IT07",
    "https://www.moneycontrol.com/india/stockpricequote/engineering-construction/bajelprojects/BP06",
    "https://www.moneycontrol.com/india/stockpricequote/vanaspatioils/bclindustries/BC06",
    "https://www.moneycontrol.com/india/stockpricequote/aerospacedefence/bharatelectronics/BE03",
    "https://www.moneycontrol.com/india/stockpricequote/power-generationdistribution/indianrenewableenergydevelopmentagency/IREDAL",
    "https://www.moneycontrol.com/india/stockpricequote/engineering-heavy/irconinternational/II07",
    "https://www.moneycontrol.com/india/stockpricequote/finance-others/jiofinancialservices/JFS",
    "https://www.moneycontrol.com/india/stockpricequote/hospitalhealthcare-services/mediassisthealthcareservices/MAH01",
    "https://www.moneycontrol.com/india/stockpricequote/finance-term-lending-institutions/rec/REC02",
    "https://www.moneycontrol.com/india/stockpricequote/food-processing/sarveshwarfoods/SF29",
    "https://www.moneycontrol.com/india/stockpricequote/banks-private-sector/yesbank/YB",
    "https://www.moneycontrol.com/india/stockpricequote/finance-investments/blb/BLB",
]

for url in url_list:
    ws.append(main(url))

wb.save('data.xlsx')